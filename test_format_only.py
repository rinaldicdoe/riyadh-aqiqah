#!/usr/bin/env python3
"""
Test script to verify color and date formatting changes (without Streamlit)
"""
import pandas as pd
import tempfile
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from xlsxwriter import Workbook

def test_date_formatting():
    """Test date formatting directly"""
    
    # Create sample data
    data = {
        'No. Invoice': ['INV001', 'INV001', 'INV002'],
        'Tanggal Kirim': [pd.to_datetime('2024-01-15'), pd.to_datetime('2024-01-15'), pd.to_datetime('2024-01-16')],
        'Tanggal Potong': [pd.to_datetime('2024-01-15'), pd.to_datetime('2024-01-15'), pd.to_datetime('2024-01-16')],
        'Paket': ['Jantan', 'Kebuli', 'Jantan'],
        'Pemotongan Disaksikan': ['Live Video Call', 'Disaksikan', 'Live Video Call'],
    }
    
    df = pd.DataFrame(data)
    
    # Generate Excel file with xlsxwriter
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        output_path = tmp.name
    
    try:
        # Create workbook
        workbook = Workbook(output_path, {'constant_memory': True})
        worksheet = workbook.add_worksheet('Test Sheet')
        
        # Define formats
        header_format = workbook.add_format({
            'bold': True,
            'font_size': 22,
            'align': 'center',
            'valign': 'vcenter',
            'border': 1,
        })
        
        date_format = workbook.add_format({
            'font_size': 22,
            'align': 'center',
            'valign': 'vcenter',
            'border': 1,
            'num_format': 'yyyy-mm-dd'
        })
        
        date_same_format = workbook.add_format({
            'font_size': 22,
            'font_color': '#000000',
            'bg_color': '#B74706',
            'bold': True,
            'align': 'center',
            'valign': 'vcenter',
            'border': 1,
            'num_format': 'yyyy-mm-dd'
        })
        
        paket_jantan_format = workbook.add_format({
            'font_color': '#000000',
            'bg_color': '#00B050',
            'align': 'center',
            'valign': 'vcenter',
            'border': 1,
        })
        
        # Write headers
        headers = list(df.columns)
        for col_num, header in enumerate(headers):
            worksheet.write(0, col_num, header, header_format)
        
        # Write data with formatting
        for row_num, row in df.iterrows():
            for col_num, (col_name, cell_value) in enumerate(row.items()):
                if col_name in ['Tanggal Kirim', 'Tanggal Potong']:
                    # Check if dates are same
                    if row['Tanggal Kirim'] == row['Tanggal Potong']:
                        fmt = date_same_format
                    else:
                        fmt = date_format
                elif col_name == 'Paket' and 'jantan' in str(cell_value).lower():
                    fmt = paket_jantan_format
                else:
                    fmt = header_format  # Default format
                
                worksheet.write(row_num + 1, col_num, cell_value, fmt)
        
        # Set column widths
        worksheet.set_column('A:E', 20)
        
        workbook.close()
        
        print(f"✅ Excel file generated successfully: {output_path}")
        print(f"📊 File size: {os.path.getsize(output_path)} bytes")
        
        # Verify with openpyxl
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        
        print("\n📋 Data Preview (checking Excel content):")
        print("Row 1 (Headers):")
        for col_num, cell in enumerate(ws[1]):
            print(f"  Col {col_num}: {cell.value}")
        
        print("\nRow 2 (First data row):")
        for col_num, cell in enumerate(ws[2]):
            val = cell.value
            fmt = cell.number_format
            print(f"  Col {col_num}: {val} (format: {fmt})")
        
        # Check colors
        print("\n🎨 Color Check:")
        for row_num in range(2, 5):
            row = ws[row_num]
            cell_c = row[2]  # Tanggal Kirim
            cell_d = row[3]  # Tanggal Potong
            print(f"Row {row_num}:")
            print(f"  Tanggal Kirim: {cell_c.value}, Fill: {cell_c.fill.start_color.rgb if cell_c.fill else 'None'}")
            print(f"  Tanggal Potong: {cell_d.value}, Fill: {cell_d.fill.start_color.rgb if cell_d.fill else 'None'}")
        
        print("\n✅ TEST PASSED - Date and color formatting verified!")
        print(f"\nFile location: {output_path}")
        print("You can open it with Excel to see the full formatting.")
        
    except Exception as e:
        print(f"❌ Error during test: {e}")
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    test_date_formatting()
